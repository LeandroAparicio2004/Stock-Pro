from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

HEADER_COLOR = "1B4F72"
ALT_ROW      = "EBF5FB"


def export_inventario_excel(productos, empresa_nombre, output_path):
    wb = Workbook()

    # ── HOJA 1: Vista bonita ──────────────────────────────
    ws = wb.active
    ws.title = "Inventario"

    ws.merge_cells("A1:J1")
    ws["A1"] = f"Inventario — {empresa_nombre}"
    ws["A1"].font  = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill  = PatternFill("solid", fgColor=HEADER_COLOR)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(italic=True, size=9, color="666666")

    headers = ["Código","Nombre","Categoría","Unidad",
               "Stock Actual","Stock Mínimo","Estado",
               "P. Costo","P. Venta","Valor Stock"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font  = Font(bold=True, color="FFFFFF")
        cell.fill  = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center")

    for i, p in enumerate(productos):
        row    = 4 + i
        estado = "⚠ BAJO" if p["stock_actual"] <= p["stock_minimo"] else "OK"
        valor  = p["stock_actual"] * p["precio_costo"]
        data   = [p["codigo"], p["nombre"], p.get("categoria_nombre",""),
                  p["unidad"], p["stock_actual"], p["stock_minimo"], estado,
                  p["precio_costo"], p["precio_venta"], valor]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(horizontal="left" if col in (2,3) else "center")
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ALT_ROW)
            if col in (8, 9, 10):
                cell.number_format = '"$"#,##0.00'
            if col == 7:
                cell.font = Font(color="E74C3C" if estado != "OK" else "27AE60", bold=True)

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=3, max_row=3+len(productos), min_col=1, max_col=10):
        for cell in row:
            cell.border = border

    for i, w in enumerate([12,30,18,10,13,13,10,12,12,14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    last = 4 + len(productos)
    ws.cell(row=last, column=1, value="TOTAL").font = Font(bold=True)
    total_val = sum(p["stock_actual"] * p["precio_costo"] for p in productos)
    c = ws.cell(row=last, column=10, value=total_val)
    c.font = Font(bold=True, color=HEADER_COLOR)
    c.number_format = '"$"#,##0.00'

    # ── HOJA 2: Para reimportar ───────────────────────────
    ws2 = wb.create_sheet("Para Importar")

    # Instrucciones
    ws2.merge_cells("A1:I1")
    ws2["A1"] = "⚠ Esta hoja sirve para actualizar productos masivamente. Modificá los valores y usá Importar Excel."
    ws2["A1"].font = Font(italic=True, size=10, color="856404")
    ws2["A1"].fill = PatternFill("solid", fgColor="FFF3CD")
    ws2["A1"].alignment = Alignment(horizontal="center")

    # Encabezados exactos que espera el importador
    imp_headers = ["codigo","nombre","descripcion","categoria",
                   "precio_costo","precio_venta",
                   "stock_actual","stock_minimo","unidad"]
    for col, h in enumerate(imp_headers, 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font  = Font(bold=True, color="FFFFFF")
        cell.fill  = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center")

    for i, p in enumerate(productos):
        row = 3 + i
        data = [
            p["codigo"],
            p["nombre"],
            p.get("descripcion",""),
            p.get("categoria_nombre","General"),
            p["precio_costo"],
            p["precio_venta"],
            p["stock_actual"],
            p["stock_minimo"],
            p["unidad"]
        ]
        for col, val in enumerate(data, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(horizontal="left" if col in (2,3,4) else "center")
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ALT_ROW)

    for i, w in enumerate([12,28,20,15,13,13,13,13,10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    return output_path


def export_movimientos_excel(movimientos, empresa_nombre, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"Movimientos — {empresa_nombre}"
    ws["A1"].font  = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill  = PatternFill("solid", fgColor=HEADER_COLOR)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["Fecha","Código","Producto","Tipo","Cantidad",
               "Stock Antes","Stock Después","Motivo"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font  = Font(bold=True, color="FFFFFF")
        cell.fill  = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center")

    tipo_color = {"entrada": "27AE60", "salida": "E74C3C", "ajuste": "F39C12"}
    for i, m in enumerate(movimientos):
        row  = 3 + i
        data = [m["fecha"], m["producto_codigo"], m["producto_nombre"],
                m["tipo"].upper(), m["cantidad"],
                m["stock_antes"], m["stock_despues"], m.get("motivo","")]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(horizontal="left" if col == 3 else "center")
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ALT_ROW)
            if col == 4:
                cell.font = Font(color=tipo_color.get(m["tipo"], "000000"), bold=True)

    for i, w in enumerate([18,12,30,10,10,12,14,25], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    return output_path


def export_inventario_pdf(productos, empresa_nombre, output_path):
    doc = SimpleDocTemplate(output_path, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles  = getSampleStyleSheet()
    story   = []

    title_style = ParagraphStyle("title", fontSize=16, fontName="Helvetica-Bold",
                                  textColor=colors.HexColor("#1B4F72"), alignment=TA_CENTER)
    sub_style   = ParagraphStyle("sub",   fontSize=9,  fontName="Helvetica",
                                  textColor=colors.grey, alignment=TA_CENTER)

    story.append(Paragraph(f"Inventario — {empresa_nombre}", title_style))
    story.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", sub_style))
    story.append(Spacer(1, 0.4*cm))

    headers    = ["Código","Nombre","Categoría","Stock","Mín.","Estado","P.Costo","P.Venta"]
    table_data = [headers]
    for p in productos:
        estado = "⚠ BAJO" if p["stock_actual"] <= p["stock_minimo"] else "✓ OK"
        table_data.append([
            p["codigo"], p["nombre"][:28], p.get("categoria_nombre","")[:15],
            str(p["stock_actual"]), str(p["stock_minimo"]), estado,
            f"${p['precio_costo']:.2f}", f"${p['precio_venta']:.2f}"
        ])

    col_widths = [2.5*cm,7*cm,4*cm,1.8*cm,1.8*cm,2.2*cm,2.5*cm,2.5*cm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    style = TableStyle([
        ("BACKGROUND",    (0,0), (-1,0),  colors.HexColor("#1B4F72")),
        ("TEXTCOLOR",     (0,0), (-1,0),  colors.white),
        ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,-1), 8),
        ("ALIGN",         (0,0), (-1,-1), "CENTER"),
        ("ALIGN",         (1,1), (1,-1),  "LEFT"),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, colors.HexColor("#EBF5FB")]),
        ("GRID",          (0,0), (-1,-1), 0.3, colors.HexColor("#CCCCCC")),
        ("TOPPADDING",    (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ])
    for i, p in enumerate(productos, 1):
        if p["stock_actual"] <= p["stock_minimo"]:
            style.add("TEXTCOLOR", (5,i), (5,i), colors.HexColor("#E74C3C"))
        else:
            style.add("TEXTCOLOR", (5,i), (5,i), colors.HexColor("#27AE60"))
        style.add("FONTNAME", (5,i), (5,i), "Helvetica-Bold")
    t.setStyle(style)
    story.append(t)

    total_val = sum(p["stock_actual"] * p["precio_costo"] for p in productos)
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(
        f"Total productos: {len(productos)}  |  Valor inventario: ${total_val:.2f}", sub_style))
    doc.build(story)
    return output_path


def import_productos_excel(db_path, filepath):
    from openpyxl import load_workbook
    from db.database import (add_producto, update_producto, get_categorias,
                              add_categoria, get_productos)
    import re

    wb = load_workbook(filepath)

    # Buscar la hoja correcta
    if "Para Importar" in wb.sheetnames:
        ws = wb["Para Importar"]
        header_row = 2
    else:
        ws = wb.active
        header_row = 1

    # Leer encabezados y normalizar
    raw_headers = [str(c.value).strip().lower() if c.value else "" for c in ws[header_row]]

    # Mapeo flexible: acepta nombres en español, inglés, abreviaturas, con espacios, etc.
    SINONIMOS = {
        "codigo":        ["codigo","código","code","cod","id","sku","ref","referencia","art","articulo","artículo"],
        "nombre":        ["nombre","name","producto","descripcion corta","desc corta","item","articulo","artículo","denominacion","denominación"],
        "descripcion":   ["descripcion","descripción","description","detalle","obs","observacion","observación","nota","notas"],
        "categoria":     ["categoria","categoría","category","rubro","tipo","familia","grupo","departamento"],
        "precio_costo":  ["precio_costo","precio costo","costo","cost","p.costo","pcosto","precio de costo","valor costo","compra","precio compra"],
        "precio_venta":  ["precio_venta","precio venta","venta","price","p.venta","pventa","precio de venta","valor venta","precio público","precio publico"],
        "stock_actual":  ["stock_actual","stock actual","stock","cantidad","qty","quantity","existencia","existencias","inventario","disponible","cant"],
        "stock_minimo":  ["stock_minimo","stock mínimo","stock minimo","minimo","mínimo","min","stock min","cantidad minima","cantidad mínima"],
        "unidad":        ["unidad","unidades","unit","um","u/m","medida","und","unid"],
    }

    def detectar_campo(header_raw):
        h = header_raw.lower().strip()
        h = re.sub(r'[_\-\.]', ' ', h).strip()
        for campo, sinonimos in SINONIMOS.items():
            for s in sinonimos:
                s2 = re.sub(r'[_\-\.]', ' ', s).strip()
                if h == s2 or h.startswith(s2):
                    return campo
        return None

    # Construir mapa: índice de columna → campo interno
    col_map = {}
    for idx, h in enumerate(raw_headers):
        campo = detectar_campo(h)
        if campo and campo not in col_map:
            col_map[campo] = idx

    if "nombre" not in col_map:
        return False, (
            "No encontré ninguna columna de nombre/producto.\n"
            "Asegurate de que el Excel tenga al menos una columna con el nombre del producto."
        )

    cats      = {c["nombre"].lower(): c["id"] for c in get_categorias(db_path)}
    existentes = {p["codigo"]: p["id"] for p in get_productos(db_path, solo_activos=False)}

    creados      = 0
    actualizados = 0
    omitidos     = 0
    contador_cod = 1  # Para generar códigos automáticos

    def get_val(row, campo, default=""):
        if campo in col_map:
            v = row[col_map[campo]]
            return v if v is not None else default
        return default

    def gen_codigo():
        nonlocal contador_cod
        while True:
            cod = f"IMP{contador_cod:04d}"
            if cod not in existentes:
                return cod
            contador_cod += 1

    for i, row in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1):

        if not any(row):
            continue

        nombre = str(get_val(row, "nombre", "")).strip()
        if not nombre or nombre.lower() in ("none", "nan", "n/a", "-", ""):
            omitidos += 1
            continue

        # Código: si está vacío o es None, generar uno automático
        codigo = str(get_val(row, "codigo", "")).strip()
        if not codigo or codigo.lower() in ("none", "nan", "n/a", "-", ""):
            codigo = gen_codigo()
            contador_cod += 1

        # Categoría
        cat_nombre = str(get_val(row, "categoria", "General")).strip()
        if not cat_nombre or cat_nombre.lower() in ("none","nan",""):
            cat_nombre = "General"
        if cat_nombre.lower() not in cats:
            add_categoria(db_path, cat_nombre)
            cats = {c["nombre"].lower(): c["id"] for c in get_categorias(db_path)}
        cat_id = cats.get(cat_nombre.lower(), cats.get("general"))

        # Numéricos — tolerante a texto, comas, símbolos de moneda
        def limpiar_num(v, entero=False):
            try:
                s = str(v).replace("$","").replace(",",".").strip()
                s = re.sub(r'[^\d\.]', '', s)
                return int(float(s)) if entero else round(float(s), 2)
            except:
                return 0

        pc     = limpiar_num(get_val(row, "precio_costo", 0))
        pv     = limpiar_num(get_val(row, "precio_venta",  0))
        stock  = limpiar_num(get_val(row, "stock_actual",  0), entero=True)
        minimo = limpiar_num(get_val(row, "stock_minimo",  0), entero=True)
        desc   = str(get_val(row, "descripcion", "")).strip()
        unidad = str(get_val(row, "unidad", "unidad")).strip()
        if not unidad or unidad.lower() in ("none","nan",""):
            unidad = "unidad"

        if codigo in existentes:
            update_producto(db_path, existentes[codigo], codigo, nombre,
                            desc, cat_id, pc, pv, minimo, unidad)
            actualizados += 1
        else:
            ok, _ = add_producto(db_path, codigo, nombre,
                                  desc, cat_id, pc, pv, stock, minimo, unidad)
            if ok:
                creados += 1
                existentes[codigo] = True
            else:
                omitidos += 1

    resumen = f"✓ Productos nuevos: {creados}\n✓ Actualizados: {actualizados}"
    if omitidos:
        resumen += f"\n⚠ Omitidos (sin nombre): {omitidos}"
    return True, resumen